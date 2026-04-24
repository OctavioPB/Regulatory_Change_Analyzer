"""Tests for the export service (PDF + Excel generation)."""

import uuid
from datetime import datetime, timezone
from unittest.mock import MagicMock

import pytest

from src.models.impact import ApprovalStatus, ImpactAlert, ImpactItem, Severity
from src.services.export_service import (
    export_alert_pdf,
    export_alert_xlsx,
    export_alerts_xlsx,
)


# ── Fixtures ──────────────────────────────────────────────────────────────────

def _make_item(severity: Severity, status: ApprovalStatus = ApprovalStatus.pending) -> MagicMock:
    item = MagicMock(spec=ImpactItem)
    item.id = uuid.uuid4()
    item.severity = severity
    item.approval_status = status
    item.impact_type = "contract"
    item.affected_name = "Master Loan Agreement"
    item.affected_ref = "Clause 4.2"
    item.suggestion = "Update the counterparty limit from 20% to 15% in this clause."
    item.reviewer_notes = None
    item.reviewed_at = None
    return item


def _make_alert(n_items: int = 3) -> MagicMock:
    alert = MagicMock(spec=ImpactAlert)
    alert.id = uuid.uuid4()
    alert.document_id = uuid.uuid4()
    alert.title = "Circular CNBV 10/2025: counterparty limit reduced"
    alert.is_read = False
    alert.created_at = datetime.now(timezone.utc)
    alert.items = [
        _make_item(Severity.high),
        _make_item(Severity.medium, ApprovalStatus.approved),
        _make_item(Severity.low),
    ][:n_items]
    return alert


# ── Excel ─────────────────────────────────────────────────────────────────────

def test_export_alert_xlsx_returns_bytes():
    alert = _make_alert()
    data = export_alert_xlsx(alert)
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_export_alert_xlsx_is_valid_xlsx():
    from openpyxl import load_workbook
    import io
    alert = _make_alert()
    data = export_alert_xlsx(alert)
    wb = load_workbook(io.BytesIO(data))
    assert "Summary" in wb.sheetnames
    assert "Impact Items" in wb.sheetnames


def test_export_alert_xlsx_items_sheet_has_correct_row_count():
    from openpyxl import load_workbook
    import io
    alert = _make_alert(n_items=3)
    data = export_alert_xlsx(alert)
    wb = load_workbook(io.BytesIO(data))
    ws = wb["Impact Items"]
    # 1 header row + 3 item rows
    assert ws.max_row == 4


def test_export_alert_xlsx_empty_items():
    alert = _make_alert(n_items=0)
    alert.items = []
    data = export_alert_xlsx(alert)
    assert len(data) > 0


def test_export_alerts_xlsx_multi_alert():
    from openpyxl import load_workbook
    import io
    alerts = [_make_alert(), _make_alert()]
    data = export_alerts_xlsx(alerts)
    wb = load_workbook(io.BytesIO(data))
    assert "All Alerts" in wb.sheetnames
    ws = wb["All Alerts"]
    # 1 header + 2 alerts
    assert ws.max_row == 3


# ── PDF ───────────────────────────────────────────────────────────────────────

def test_export_alert_pdf_returns_bytes():
    alert = _make_alert()
    data = export_alert_pdf(alert)
    assert isinstance(data, bytes)
    assert len(data) > 0


def test_export_alert_pdf_starts_with_pdf_magic():
    alert = _make_alert()
    data = export_alert_pdf(alert)
    # PDF files start with %PDF-
    assert data[:4] == b"%PDF"


def test_export_alert_pdf_empty_items():
    alert = _make_alert(n_items=0)
    alert.items = []
    data = export_alert_pdf(alert)
    assert data[:4] == b"%PDF"


def test_export_alert_pdf_item_with_reviewer_notes():
    alert = _make_alert(n_items=1)
    alert.items[0].reviewer_notes = "Verified with legal team."
    alert.items[0].reviewed_at = datetime.now(timezone.utc)
    data = export_alert_pdf(alert)
    assert data[:4] == b"%PDF"
