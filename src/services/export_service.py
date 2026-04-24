"""Export service — Sprint 5.

Generates PDF and Excel reports from ImpactAlerts for audit purposes.
Both formats include: document metadata, detected changes, impact items,
and reviewer decisions.
"""
import io
import logging
from datetime import datetime, timezone

from src.models.impact import ImpactAlert, ImpactItem, Severity

logger = logging.getLogger(__name__)

# ── Severity colours (shared) ─────────────────────────────────────────────────

_SEV_COLOUR = {
    Severity.high: ("#dc2626", "FF4444"),   # (hex for reportlab, ARGB for openpyxl)
    Severity.medium: ("#d97706", "FFAA00"),
    Severity.low: ("#16a34a", "22AA44"),
}


# ── Excel export ──────────────────────────────────────────────────────────────

def export_alert_xlsx(alert: ImpactAlert) -> bytes:
    """Generate an Excel workbook for a single ImpactAlert.

    Sheets:
        Summary     — alert metadata + item counts by severity
        Impact Items — one row per ImpactItem with all fields

    Args:
        alert: Fully-loaded ImpactAlert ORM object (items eagerly loaded).

    Returns:
        Raw bytes of the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws_sum = wb.active
    assert ws_sum is not None
    ws_sum.title = "Summary"

    header_font = Font(bold=True, size=12)
    ws_sum.append(["Regulatory Change Analyzer — Impact Report"])
    ws_sum["A1"].font = Font(bold=True, size=14)
    ws_sum.append([])
    ws_sum.append(["Alert title", alert.title])
    ws_sum.append(["Alert ID", str(alert.id)])
    ws_sum.append(["Document ID", str(alert.document_id)])
    ws_sum.append(["Generated at", datetime.now(timezone.utc).isoformat(timespec="seconds")])
    ws_sum.append([])

    ws_sum.append(["Severity breakdown"])
    ws_sum["A8"].font = header_font
    counts = {"high": 0, "medium": 0, "low": 0}
    for item in alert.items:
        counts[item.severity.value] += 1
    for sev, n in counts.items():
        _, argb = _SEV_COLOUR[Severity(sev)]
        row_idx = ws_sum.max_row + 1
        ws_sum.append([sev.upper(), n])
        ws_sum.cell(row_idx, 1).fill = PatternFill("solid", fgColor=argb)
        ws_sum.cell(row_idx, 1).font = Font(bold=True, color="FFFFFF")

    ws_sum.column_dimensions["A"].width = 22
    ws_sum.column_dimensions["B"].width = 50

    # ── Sheet 2: Impact Items ─────────────────────────────────────────────────
    ws_items = wb.create_sheet("Impact Items")

    col_headers = [
        "Severity", "Affected Name", "Affected Ref",
        "Type", "Suggestion", "Status", "Reviewer Notes", "Reviewed At",
    ]
    ws_items.append(col_headers)
    for col_idx, header in enumerate(col_headers, 1):
        cell = ws_items.cell(1, col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.alignment = Alignment(horizontal="center")

    for item in alert.items:
        _, argb = _SEV_COLOUR[item.severity]
        row = [
            item.severity.value.upper(),
            item.affected_name,
            item.affected_ref or "",
            item.impact_type,
            item.suggestion,
            item.approval_status.value,
            item.reviewer_notes or "",
            item.reviewed_at.isoformat(timespec="seconds") if item.reviewed_at else "",
        ]
        ws_items.append(row)
        row_idx = ws_items.max_row
        ws_items.cell(row_idx, 1).fill = PatternFill("solid", fgColor=argb)
        ws_items.cell(row_idx, 1).font = Font(bold=True, color="FFFFFF")

    # Column widths
    widths = [12, 35, 20, 12, 70, 12, 40, 22]
    for i, w in enumerate(widths, 1):
        ws_items.column_dimensions[get_column_letter(i)].width = w
    for row in ws_items.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    buf = io.BytesIO()
    wb.save(buf)
    logger.info("Excel report generated for alert %s (%d items)", alert.id, len(alert.items))
    return buf.getvalue()


def export_alerts_xlsx(alerts: list[ImpactAlert]) -> bytes:
    """Generate a multi-sheet Excel workbook for a list of alerts.

    One sheet per alert, plus a master summary sheet.

    Args:
        alerts: List of fully-loaded ImpactAlert objects.

    Returns:
        Raw bytes of the .xlsx file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws_master = wb.active
    assert ws_master is not None
    ws_master.title = "All Alerts"

    master_headers = ["Title", "Document ID", "Items", "High", "Medium", "Low", "Created At"]
    ws_master.append(master_headers)
    for col_idx, h in enumerate(master_headers, 1):
        cell = ws_master.cell(1, col_idx)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E3A5F")

    for alert in alerts:
        counts = {"high": 0, "medium": 0, "low": 0}
        for item in alert.items:
            counts[item.severity.value] += 1
        ws_master.append([
            alert.title,
            str(alert.document_id),
            len(alert.items),
            counts["high"],
            counts["medium"],
            counts["low"],
            alert.created_at.isoformat(timespec="seconds"),
        ])

    for i, w in enumerate([60, 38, 8, 8, 8, 8, 22], 1):
        ws_master.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    logger.info("Multi-alert Excel report generated (%d alerts)", len(alerts))
    return buf.getvalue()


# ── PDF export ────────────────────────────────────────────────────────────────

def export_alert_pdf(alert: ImpactAlert) -> bytes:
    """Generate a PDF report for a single ImpactAlert.

    Args:
        alert: Fully-loaded ImpactAlert ORM object (items eagerly loaded).

    Returns:
        Raw bytes of the .pdf file.
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        HRFlowable, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=styles["Heading1"], fontSize=16, spaceAfter=6)
    h2 = ParagraphStyle("h2", parent=styles["Heading2"], fontSize=12, spaceAfter=4)
    normal = ParagraphStyle("normal", parent=styles["Normal"], fontSize=9, spaceAfter=4)
    small = ParagraphStyle("small", parent=styles["Normal"], fontSize=8, textColor=colors.gray)

    _SEV_RL = {
        Severity.high: colors.HexColor("#dc2626"),
        Severity.medium: colors.HexColor("#d97706"),
        Severity.low: colors.HexColor("#16a34a"),
    }

    story = []

    # Title
    story.append(Paragraph("Regulatory Change Analyzer", styles["Title"]))
    story.append(Paragraph("Impact Report", h1))
    story.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#1E3A5F")))
    story.append(Spacer(1, 0.3 * cm))

    # Alert metadata
    meta_data = [
        ["Alert:", alert.title],
        ["Alert ID:", str(alert.id)],
        ["Document ID:", str(alert.document_id)],
        ["Generated:", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")],
    ]
    meta_table = Table(meta_data, colWidths=[3 * cm, 14 * cm])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#1E3A5F")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    story.append(meta_table)
    story.append(Spacer(1, 0.5 * cm))

    # Severity summary
    story.append(Paragraph("Severity Summary", h2))
    counts = {"high": 0, "medium": 0, "low": 0}
    for item in alert.items:
        counts[item.severity.value] += 1
    sev_data = [["Severity", "Count"]] + [
        [sev.upper(), str(n)] for sev, n in counts.items() if n > 0
    ]
    if len(sev_data) > 1:
        sev_table = Table(sev_data, colWidths=[4 * cm, 3 * cm])
        sev_style = [
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E3A5F")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (1, 0), (1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ]
        for i, (sev, _) in enumerate(counts.items(), 1):
            if counts[sev] > 0:
                sev_style.append(("TEXTCOLOR", (0, i), (0, i), _SEV_RL[Severity(sev)]))
                sev_style.append(("FONTNAME", (0, i), (0, i), "Helvetica-Bold"))
        sev_table.setStyle(TableStyle(sev_style))
        story.append(sev_table)
    story.append(Spacer(1, 0.5 * cm))

    # Impact items
    story.append(Paragraph(f"Impact Items ({len(alert.items)})", h2))
    story.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey))
    story.append(Spacer(1, 0.2 * cm))

    for idx, item in enumerate(alert.items, 1):
        sev_color = _SEV_RL[item.severity]

        story.append(
            Paragraph(
                f'<font color="#{_SEV_RL[item.severity].hexval()[2:]}" size="10"><b>'
                f"[{item.severity.value.upper()}]</b></font> "
                f"<b>{item.affected_name}</b>"
                + (f" · <i>{item.affected_ref}</i>" if item.affected_ref else ""),
                normal,
            )
        )
        story.append(Paragraph(f"<b>Type:</b> {item.impact_type}", small))
        story.append(Paragraph(f"<b>Status:</b> {item.approval_status.value}", small))
        story.append(Paragraph(f"<b>Suggestion:</b> {item.suggestion}", normal))
        if item.reviewer_notes:
            story.append(Paragraph(f"<b>Reviewer notes:</b> {item.reviewer_notes}", small))
        if idx < len(alert.items):
            story.append(HRFlowable(width="100%", thickness=0.3, color=colors.HexColor("#e5e7eb")))
            story.append(Spacer(1, 0.2 * cm))

    doc.build(story)
    logger.info("PDF report generated for alert %s (%d items)", alert.id, len(alert.items))
    return buf.getvalue()
